from typing import List, Any, Optional, Tuple
from dataclasses import dataclass, field
import openpyxl
from fastapi import UploadFile

@dataclass
class MergeRange:
    start_row: int
    start_col: int
    end_row: int
    end_col: int

@dataclass
class SheetGrid:
    sheet_name: str
    ref_range: str
    n_rows: int
    n_cols: int
    grid: List[List[Any]]
    merges: List[MergeRange] = field(default_factory=list)

class DataLoader:
    @staticmethod
    def load_xlsx(file_content: bytes) -> SheetGrid:
        from io import BytesIO
        wb = openpyxl.load_workbook(filename=BytesIO(file_content), data_only=True)
        ws = wb.active # Default to first sheet
        
        # Determine actual dimensions
        # openpyxl max_row/max_column can be unreliable if there is formatting but no data
        # We will use calculate_dimension or strict iteration
        # For this implementation, let's use list(ws.values) but we need to handle merges manually
        # So we iterate cells.
        
        data = []
        max_r = ws.max_row
        max_c = ws.max_column
        
        # Read grid
        # NOTE: openpyxl is 1-indexed, our grid is 0-indexed
        for r in range(1, max_r + 1):
            row_data = []
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                row_data.append(cell.value)
            data.append(row_data)

        # Capture merges
        merges = []
        for merge in ws.merged_cells.ranges:
            # openpyxl range: min_row, min_col, max_row, max_col (1-indexed)
            # Convert to 0-indexed
            merges.append(MergeRange(
                start_row=merge.min_row - 1,
                start_col=merge.min_col - 1,
                end_row=merge.max_row - 1,
                end_col=merge.max_col - 1
            ))
            
            # CLEAR content of merged cells except top-left
            # openpyxl might give value only for top-left, but let's ensure consistency
            # If top-left is (mr, mc), then any other cell in range should be None in our grid
            for cx in range(merge.min_col, merge.max_col + 1):
                for rx in range(merge.min_row, merge.max_row + 1):
                    if rx == merge.min_row and cx == merge.min_col:
                        continue
                    # 0-indexed access
                    data[rx-1][cx-1] = None

        return SheetGrid(
            sheet_name=ws.title,
            ref_range=ws.dimensions,
            n_rows=max_r,
            n_cols=max_c,
            grid=data,
            merges=merges
        )

    @staticmethod
    def load_xls(file_content: bytes) -> SheetGrid:
        # TODO: Implement XLRD or Converter
        # For now raise NotImplementedError or return empty
        raise NotImplementedError("XLS support not yet implemented. Please convert to XLSX.")
