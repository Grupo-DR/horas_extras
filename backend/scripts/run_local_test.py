import sys
import os
import asyncio
import openpyxl

# Add backend to sys.path
sys.path.append(os.path.join(os.path.dirname(__file__), ".."))

from app.modules.loader import DataLoader
from app.modules.extractor import Extractor

def create_sample_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BM Detalhado"
    
    # Metadata
    ws["B2"] = "Contrato N.º"
    ws["C2"] = "4600012092"
    ws["B3"] = "Data:"
    ws["C3"] = "12/01/2025"
    
    # Header
    headers = ["ITEM", "DESCRIÇÃO DOS SERVIÇOS", "UNID", "PREVISTO CONTRATO", "SALDO A MEDIR", "EXEC %", "VALOR MEDIDO"]
    # Row 10
    for i, h in enumerate(headers, start=2): # Start at Column B
        ws.cell(row=10, column=i, value=h)
        
    # Items
    items = [
        ("1.1", "Mobilização", "UN", 10000.0, 0.0, 100, 10000.0),
        ("1.2", "Escavação", "m3", 50000.0, 25000.0, 50, 25000.0),
        ("2.1", "Concreto", "m3", 100000.0, 100000.0, 0, 0.0),
    ]
    
    start_row = 11
    for i, item in enumerate(items):
        r = start_row + i
        ws.cell(row=r, column=2, value=item[0]) # Item
        ws.cell(row=r, column=3, value=item[1]) # Desc
        ws.cell(row=r, column=4, value=item[2]) # Unid
        ws.cell(row=r, column=5, value=item[3]) # Previsto
        ws.cell(row=r, column=6, value=item[4]) # Saldo
        ws.cell(row=r, column=7, value=item[5]) # Exec
        ws.cell(row=r, column=8, value=item[6]) # Medido

    # Total
    ws.cell(row=start_row + len(items) + 2, column=3, value="TOTAL GERAL")
    
    wb.save(filename)
    print(f"Created {filename}")

def main():
    test_file = "test_bm.xlsx"
    create_sample_xlsx(test_file)
    
    with open(test_file, "rb") as f:
        content = f.read()
    
    # Run Loader
    print("Loading...")
    grid = DataLoader.load_xlsx(content)
    print(f"Loaded Grid: {grid.n_rows}x{grid.n_cols}")
    
    # Run Extractor
    print("Extracting...")
    extractor = Extractor(grid)
    bm = extractor.extract()
    
    print("\n--- Parsing Result ---")
    print(f"Layout Confidence: {bm.overall_confidence}")
    print(f"Found {len(bm.items)} items")
    for item in bm.items:
        print(f"[{item.item_code.value}] {item.descricao.value} | Exec: {item.exec_percent.value}%")
    
    print("\nWarning:", bm.warnings)
    
    # Cleanup
    if os.path.exists(test_file):
        os.remove(test_file)

if __name__ == "__main__":
    main()
